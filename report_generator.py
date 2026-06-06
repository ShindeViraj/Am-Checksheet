import io
import os
import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Load the JSON data once at module startup
DATA_FILE = os.path.join(os.path.dirname(__file__), 'storage', 'machine_data.json')
with open(DATA_FILE, 'r', encoding='utf-8') as f:
    MACHINE_TEMPLATES = json.load(f)


def find_template_data(machine_id):
    """Fuzzy search for the machine ID in the JSON keys"""
    machine_id = machine_id.strip().lower()
    for key, data in MACHINE_TEMPLATES.items():
        if machine_id in key.lower():
            return data
    return None


# ---------------------------------------------------------------------------
# Column layout:
#   A (1) : S.No.
#   B (2) : SQM
#   C (3) : Check Point
#   D (4) : Spec. Type (A/V)
#   E (5) : Shift
#   Then for each day d (1‑31):  mark + time columns
# ---------------------------------------------------------------------------
HEADER_COLS = 5
FIRST_DAY_COL = 6          # column where Day 1 mark starts

def _day_mark_col(day):
    return FIRST_DAY_COL + (day - 1) * 2

def _day_time_col(day):
    return FIRST_DAY_COL + (day - 1) * 2 + 1

LAST_DATA_COL = _day_time_col(31)  # 67


def _estimate_row_height(text, col_width_chars=83, font_size=14):
    """Estimate the row height needed for wrapped text in a merged cell.
    Returns height in points for the LAST row of a 3-row merged block.
    The first two rows use default height (~15pt each = 30pt).
    """
    if not text:
        return 15
    # Approximate chars per line at font_size 14 in a column of width 83
    chars_per_line = int(col_width_chars * 0.95)
    lines = 0
    for paragraph in str(text).split('\n'):
        paragraph = paragraph.strip()
        if not paragraph:
            continue          # skip empty lines (collapsed \n\n)
        else:
            lines += max(1, -(-len(paragraph) // chars_per_line))  # ceil division
    # Total height needed for all lines at ~20pt per line (generous for zoom)
    total_height = lines * 20
    # First two rows provide ~30pt, so last row needs the remainder
    remaining = max(18, total_height - 30)
    return remaining


def generate_report(machine_id, month_str, checkpoints_data, start_date_str='2026-05-01', end_date_str='2026-05-31'):
    """
    Build a fresh Excel workbook matching the user's edited template layout.
    """
    template_data = find_template_data(machine_id)
    if not template_data:
        template_data = {
            'title': 'Autonomous Maintenance (Daily) Check List',
            'line_name': 'Line:',
            'machine_name': f'Machine Name: {machine_id}',
            'types_text': 'Type: Safety (S) Poka-Yoke (Q) Daily PM (M)',
            'checkpoints': []
        }

    from datetime import datetime, timedelta
    dates_list = []
    if start_date_str and end_date_str:
        s = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        e = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        curr = s
        while curr <= e:
            dates_list.append(curr)
            curr += timedelta(days=1)
    if not dates_list:
        dates_list = [datetime.now().date()] # fallback

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    # ── Styles ──────────────────────────────────────────────────────────────
    FONT_16B  = Font(name='Calibri', size=16, bold=True)
    FONT_14B  = Font(name='Calibri', size=14, bold=True)
    FONT_14   = Font(name='Calibri', size=14)

    center_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_wrap   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    center_nw   = Alignment(horizontal='center', vertical='center', wrap_text=False)

    thin_side   = Side(border_style='thin',   color='000000')
    medium_side = Side(border_style='medium', color='000000')

    thin_border   = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    medium_border = Border(left=medium_side, right=medium_side, top=medium_side, bottom=medium_side)

    hdr_fill   = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    shift_b    = PatternFill(start_color='8DB4E2', end_color='8DB4E2', fill_type='solid')
    time_fill  = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    # ── Column widths ───────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 7.5      # S.No.
    ws.column_dimensions['B'].width = 10.7     # SQM
    ws.column_dimensions['C'].width = 83.5     # Check Point
    ws.column_dimensions['D'].width = 10.0     # Spec. Type
    ws.column_dimensions['E'].width = 10.6     # Shift

    for day in range(1, 32):
        mc = get_column_letter(_day_mark_col(day))
        tc = get_column_letter(_day_time_col(day))
        ws.column_dimensions[mc].width = 7.7   # ✓/✗ mark
        ws.column_dimensions[tc].width = 8.2   # T(m) time

    # ── Helper ──────────────────────────────────────────────────────────────
    def cell(r, c, val, font=FONT_14, align=center_wrap, fill=None, brd=thin_border):
        cl = ws.cell(row=r, column=c, value=val)
        if font:  cl.font = font
        if align: cl.alignment = align
        if fill:  cl.fill = fill
        if brd:   cl.border = brd
        return cl

    # ── Row 1 – Title ───────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=HEADER_COLS)
    cell(1, 1, template_data['title'], font=FONT_16B, align=center_wrap, brd=medium_border)
    for c in range(2, HEADER_COLS + 1):
        ws.cell(1, c).border = medium_border
    ws.row_dimensions[1].height = 30

    # ── Row 2 – Line / Machine Name ─────────────────────────────────────────
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
    cell(2, 1, template_data['line_name'], font=FONT_16B, align=center_wrap, brd=medium_border)
    ws.cell(2, 2).border = medium_border

    ws.merge_cells(start_row=2, start_column=3, end_row=2, end_column=HEADER_COLS)
    cell(2, 3, template_data['machine_name'], font=FONT_16B, align=center_wrap, brd=medium_border)
    for c in range(4, HEADER_COLS + 1):
        ws.cell(2, c).border = medium_border

    # Extend row 2 merge into day columns (empty, no border)
    ws.merge_cells(start_row=2, start_column=FIRST_DAY_COL,
                   end_row=2, end_column=LAST_DATA_COL)
    ws.row_dimensions[2].height = 25

    # ── Row 3 – Month / Date Range ──────────────────────────────────────────
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=HEADER_COLS)
    cell(3, 1, month_str, font=FONT_16B, align=center_wrap, brd=medium_border)
    for c in range(2, HEADER_COLS + 1):
        ws.cell(3, c).border = medium_border
    ws.row_dimensions[3].height = 26

    # ── Row 4 – Type legend only (no CLIT) ──────────────────────────────────
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=HEADER_COLS)
    cell(4, 1, template_data['types_text'], font=FONT_14B, align=center_wrap, brd=medium_border)
    for c in range(2, HEADER_COLS + 1):
        ws.cell(4, c).border = medium_border

    # Row 4 also merges into day columns with the type text extending
    ws.merge_cells(start_row=4, start_column=FIRST_DAY_COL,
                   end_row=4, end_column=LAST_DATA_COL)
    ws.row_dimensions[4].height = 29

    # ── Row 5 – Column headers ──────────────────────────────────────────────
    fixed_headers = ['S.No.', 'SQM', 'Check Point', 'Spec.\nType', 'Shift']
    for idx, hdr in enumerate(fixed_headers):
        cell(5, idx + 1, hdr, font=FONT_14B, fill=hdr_fill)

    # Day headers – merge mark + time per day
    for idx, d in enumerate(dates_list, start=1):
        mc = _day_mark_col(idx)
        tc = _day_time_col(idx)
        ws.merge_cells(start_row=5, start_column=mc, end_row=5, end_column=tc)
        cell(5, mc, d.day, font=FONT_14B, fill=hdr_fill)
        ws.cell(5, tc).border = thin_border
        ws.cell(5, tc).fill = hdr_fill
    ws.row_dimensions[5].height = 36

    # ── Row 6 – Sub-headers (✓/✗ | T(m)) ───────────────────────────────────
    for col in range(1, HEADER_COLS + 1):
        cell(6, col, None, fill=hdr_fill, font=FONT_14B)
    for idx in range(1, len(dates_list) + 1):
        cell(6, _day_mark_col(idx), '✓/✗', font=FONT_14B, fill=hdr_fill)
        cell(6, _day_time_col(idx), 'T(m)', font=FONT_14B, fill=time_fill)
    ws.row_dimensions[6].height = 54

    # ── Checkpoint rows (3 per checkpoint: Shift A / B / C) ─────────────────
    DATA_START_ROW = 7
    current_row = DATA_START_ROW

    for cp_idx, cp in enumerate(template_data['checkpoints']):
        r_a = current_row       # Shift A row
        r_b = current_row + 1   # Shift B row
        r_c = current_row + 2   # Shift C row

        # --- Merge S.No. (col A) across 3 rows – BOLD ---
        ws.merge_cells(start_row=r_a, start_column=1, end_row=r_c, end_column=1)
        cell(r_a, 1, cp['s_no'], font=FONT_14B, align=center_wrap)
        for sr in range(r_a, r_c + 1):
            ws.cell(sr, 1).border = thin_border

        # --- Merge SQM (col B) across 3 rows – BOLD ---
        ws.merge_cells(start_row=r_a, start_column=2, end_row=r_c, end_column=2)
        cell(r_a, 2, cp['sqm'], font=FONT_14B, align=center_wrap)
        for sr in range(r_a, r_c + 1):
            ws.cell(sr, 2).border = thin_border

        # --- Merge Check Point (col C) across 3 rows – BOLD, left-aligned ---
        # Collapse double newlines to single so there's no blank gap
        desc_text = cp['check_point'].replace('\n\n', '\n') if cp.get('check_point') else ''
        ws.merge_cells(start_row=r_a, start_column=3, end_row=r_c, end_column=3)
        cell(r_a, 3, desc_text, font=FONT_14B, align=left_wrap)
        for sr in range(r_a, r_c + 1):
            ws.cell(sr, 3).border = thin_border

        # --- Merge Spec Type (col D) across 3 rows ---
        ws.merge_cells(start_row=r_a, start_column=4, end_row=r_c, end_column=4)
        cell(r_a, 4, cp['spec'], font=FONT_14, align=center_wrap)
        for sr in range(r_a, r_c + 1):
            ws.cell(sr, 4).border = thin_border

        # --- Shift labels + empty day cells ---
        for shift_idx, shift_name in enumerate(['A', 'B', 'C']):
            r = current_row + shift_idx
            is_b = (shift_name == 'B')
            fill_s = shift_b if is_b else None

            cell(r, 5, shift_name, font=FONT_14, fill=fill_s)

            for idx in range(1, len(dates_list) + 1):
                mc = _day_mark_col(idx)
                tc = _day_time_col(idx)
                mark_fill = shift_b if is_b else None
                tm_fill   = shift_b if is_b else time_fill
                cell(r, mc, None, font=FONT_14, fill=mark_fill)
                cell(r, tc, None, font=FONT_14, fill=tm_fill)

        # --- Row height for the C-shift row (last of 3) ---
        last_row_h = _estimate_row_height(desc_text)
        ws.row_dimensions[r_c].height = last_row_h

        current_row += 3

    # ── Fill in database data ───────────────────────────────────────────────
    def get_shift_and_date(dt):
        """Return (shift_offset, logical_date).
        C shift (00:00‑06:59) belongs to the PREVIOUS calendar day.
        """
        mins = dt.hour * 60 + dt.minute
        if 420 <= mins < 930:       # A  07:00‑15:29
            return 0, dt.date()
        elif 930 <= mins < 1440:    # B  15:30‑23:59
            return 1, dt.date()
        else:                       # C  00:00‑06:59 → previous day
            from datetime import timedelta
            return 2, (dt - timedelta(days=1)).date()

    shift_times = {}   # (day_idx, shift) → {first, last}

    for rec in checkpoints_data:
        cp_no  = rec['checkpoint_no']
        shift, logical_date = get_shift_and_date(rec['start_time'])

        try:
            day_idx = dates_list.index(logical_date) + 1
        except ValueError:
            continue

        cp_row = DATA_START_ROW + (cp_no - 1) * 3 + shift

        if cp_row >= current_row:
            continue

        mark = '✓' if rec['checkpoint_ok'] else ('✗' if rec['checkpoint_not_ok'] else '-')
        is_b = (shift == 1)
        mark_fill = shift_b if is_b else None
        cell(cp_row, _day_mark_col(day_idx), mark, font=FONT_14B, fill=mark_fill)

        time_val = float(rec['time_taken']) if rec['time_taken'] else 0
        if time_val > 0:
            tm_fill = shift_b if is_b else time_fill
            cell(cp_row, _day_time_col(day_idx), round(time_val, 1),
                 font=FONT_14, fill=tm_fill)

        # track shift timing
        key = (day_idx, shift)
        if key not in shift_times:
            shift_times[key] = {'first': rec['start_time'], 'last': rec['end_time']}
        else:
            if rec['start_time'] < shift_times[key]['first']:
                shift_times[key]['first'] = rec['start_time']
            if rec['end_time'] and (shift_times[key]['last'] is None
                                    or rec['end_time'] > shift_times[key]['last']):
                shift_times[key]['last'] = rec['end_time']

    # ── Summary section ─────────────────────────────────────────────────────
    # Blank row after data, then:
    # Shift A: 3 rows, blank row, Shift B: 3 rows (blue), blank row, Shift C: 3 rows
    summary_start = current_row + 1
    labels = ['Checksheet Start', 'Checksheet End', 'Total Time (min)']
    shift_names_list = ['A', 'B', 'C']
    shift_offsets = [0, 4, 8]  # row offset for each shift block

    for s_idx, (shift_name, offset) in enumerate(zip(shift_names_list, shift_offsets)):
        is_b = (shift_name == 'B')
        fill = shift_b if is_b else None

        for i, label in enumerate(labels):
            r = summary_start + offset + i

            # Merge A:C for label
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
            cell(r, 1, label, font=FONT_14B, align=center_wrap, fill=fill)
            for sc in range(2, 4):
                ws.cell(sc, sc).border = thin_border  # slave cells
                ws.cell(r, sc).border = thin_border
                if fill: ws.cell(r, sc).fill = fill

            # Col D empty
            cell(r, 4, None, font=FONT_14, fill=fill)

            # Col E = Shift label
            cell(r, 5, f'Shift {shift_name}', font=FONT_14B, fill=fill)

            # Day columns – merge mark+time per day
            for idx in range(1, len(dates_list) + 1):
                mc = _day_mark_col(idx)
                tc = _day_time_col(idx)
                ws.merge_cells(start_row=r, start_column=mc, end_row=r, end_column=tc)
                cell(r, mc, None, font=FONT_14, fill=fill)
                ws.cell(r, tc).border = thin_border
                if fill:
                    ws.cell(r, tc).fill = fill

            # Set row height for summary rows
            ws.row_dimensions[r].height = 16

    # Fill summary data – show HH:MM:SS for start/end
    for (day_idx, shift), times in shift_times.items():
        if times['first'] and times['last']:
            start_str = times['first'].strftime('%H:%M:%S')
            end_str   = times['last'].strftime('%H:%M:%S')
            diff_mins = (times['last'] - times['first']).total_seconds() / 60.0

            mc = _day_mark_col(day_idx)
            s_row = summary_start + shift_offsets[shift]
            is_b = (shift == 1)
            fill = shift_b if is_b else None

            cell(s_row,     mc, start_str, font=FONT_14, fill=fill)
            cell(s_row + 1, mc, end_str,   font=FONT_14, fill=fill)
            cell(s_row + 2, mc, round(diff_mins, 1), font=FONT_14, fill=fill)

    # ── Freeze panes ────────────────────────────────────────────────────────
    ws.freeze_panes = 'F7'

    # ── Write out ───────────────────────────────────────────────────────────
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out
